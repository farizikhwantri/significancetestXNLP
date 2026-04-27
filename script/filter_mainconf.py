import re
import os
import argparse
from typing import Dict, List, Optional

import pandas as pd
import bibtexparser
from openpyxl import load_workbook

# Main-conference patterns applied to booktitle/series only
VENUE_PATTERNS = {
    "ACL": [
        r"\bproceedings of (?:the )?(?:\d{1,2}(?:st|nd|rd|th)\s)?annual meeting of the association for computational linguistics\b",
        r"\bannual meeting of the association for computational linguistics\b",
        r"\bacl\b"
    ],
    "NAACL": [
        r"\b(?:conference|meeting) of the north american chapter of the association for computational linguistics\b",
        r"\bnaacl\b"
    ],
    "AACL/IJCNLP": [
        r"\basia[- ]pacific chapter of the association for computational linguistics\b",
        r"\baacl(?:-|/)?ijcnlp\b"
    ],
    "EMNLP": [
        r"\bconference on empirical methods in natural language processing\b",
        r"\bemnlp\b"
    ],
    "EACL": [
        r"\b(?:conference|meeting) of the european chapter of the association for computational linguistics\b",
        r"\beacl\b"
    ],
    "COLING": [
        r"\binternational conference on computational linguistics\b",
        r"\bcoling\b"
    ],
    "LREC": [
        r"\binternational conference on language resources and evaluation\b",
        r"\blrec\b"
    ],
    "Findings": [
        r"\bfindings of (?:the )?(?:association for computational linguistics|acl|naacl|emnlp|eacl|ijcnlp|aacl)\b"
    ],
}

# Exclude non-main tracks by booktitle/series/url
EXCLUDE_BOOKTITLE = [
    r"\bworkshop\b",
    r"\bstudent research workshop\b",
    r"\bsystem (?:demonstrations?|demo)\b",
    r"\bdemo\b",
    r"\bshared task\b",
    r"\btutorials?\b",
    r"\bindustry track\b",
    r"\beval4nlp\b",
    r"\bblackboxnlp\b",
    r"\bwassa\b",
    r"\bwoah\b",
    r"\bimageeval\b",
    r"\bfever\b",
]

# ACL Anthology url helpers (non-id based)
URL_POSITIVE = {
    "CL": re.compile(r"/\d{4}\.cl-", re.IGNORECASE),
    "TACL": re.compile(r"/\d{4}\.tacl-", re.IGNORECASE),
    "CONF_MAIN": re.compile(r"/\d{4}\.(acl|naacl|emnlp|eacl|coling|lrec)-", re.IGNORECASE),
    "FINDINGS": re.compile(r"/\d{4}\.findings-(acl|naacl|emnlp|eacl|ijcnlp|aacl)", re.IGNORECASE),
}
URL_EXCLUDE = re.compile(r"/\d{4}\.(.*workshop|ws|demo|student|tutorial|industry|fever|blackboxnlp|wassa|eval4nlp|woah)", re.IGNORECASE)

def compile_patterns(pattern_dict):
    compiled = {}
    for venue, pats in pattern_dict.items():
        compiled[venue] = [re.compile(p, re.IGNORECASE) for p in pats]
    return compiled

def normalize(s: Optional[str]) -> str:
    s = (s or "").lower()
    s = re.sub(r"[{}]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def matches_any(text: str, regs: List[re.Pattern]) -> bool:
    return any(r.search(text) for r in regs)

def load_bib_index(bib_paths: List[str]) -> Dict[str, Dict]:
    index: Dict[str, Dict] = {}
    for p in bib_paths:
        if not os.path.exists(p):
            print(f"Warning: bib file not found: {p}")
            continue
        with open(p, "r", encoding="utf-8") as fh:
            db = bibtexparser.load(fh)
            for entry in db.entries:
                bib_id = entry.get("ID")
                if bib_id:
                    index[bib_id] = entry
    return index

def is_non_main_track(booktitle: str, series: str, url: str) -> bool:
    bt = booktitle or series
    if bt and any(re.search(p, bt, flags=re.IGNORECASE) for p in EXCLUDE_BOOKTITLE):
        return True
    if url and URL_EXCLUDE.search(url):
        return True
    return False

def detect_venue(entry: Optional[Dict], compiled) -> Optional[str]:
    if not entry:
        return None

    journal = normalize(entry.get("journal"))
    booktitle = normalize(entry.get("booktitle"))
    series = normalize(entry.get("series"))
    url = normalize(entry.get("url"))
    publisher = normalize(entry.get("publisher"))

    # Journals: journal field or ACL Anthology url only
    if journal:
        if "transactions of the association for computational linguistics" in journal or journal.strip() == "tacl":
            return "TACL"
        if journal.strip() == "computational linguistics":
            return "Computational Linguistics (CL)"
    if url:
        if URL_POSITIVE["TACL"].search(url):
            return "TACL"
        if URL_POSITIVE["CL"].search(url):
            return "Computational Linguistics (CL)"

    # Exclude workshops/etc. before matching conferences
    if is_non_main_track(booktitle, series, url):
        return None

    # Findings via booktitle/series or url
    bt = booktitle or series
    if bt and matches_any(bt, compiled["Findings"]):
        return "Findings"
    if url and URL_POSITIVE["FINDINGS"].search(url):
        return "Findings"

    # Main conferences via booktitle/series or url
    for venue in ["ACL", "NAACL", "AACL/IJCNLP", "EMNLP", "EACL", "COLING", "LREC"]:
        regs = compiled[venue]
        if bt and matches_any(bt, regs):
            return venue
    if url and URL_POSITIVE["CONF_MAIN"].search(url):
        # Try to map the url token to a specific venue if booktitle missing
        m = URL_POSITIVE["CONF_MAIN"].search(url)
        token = m.group(1).upper() if m else None
        if token == "ACL":
            return "ACL"
        if token == "NAACL":
            return "NAACL"
        if token == "EMNLP":
            return "EMNLP"
        if token == "EACL":
            return "EACL"
        if token == "COLING":
            return "COLING"
        if token == "LREC":
            return "LREC"

    return None

def load_input_table(path: str, title_col: str, id_col: str, excel_read: str = "values") -> pd.DataFrame:
    ext = os.path.splitext(path.lower())[1]
    if ext in [".xlsx", ".xls"]:
        data_only = (excel_read == "values")
        wb = load_workbook(path, data_only=data_only, read_only=True)
        ws = wb.worksheets[0]
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
        headers = [("" if c.value is None else str(c.value).strip()) for c in header_cells]
        rows = []
        for row_cells in ws.iter_rows(min_row=2, values_only=False):
            row = {}
            for j, cell in enumerate(row_cells):
                val = cell.value
                row[headers[j] if j < len(headers) else f"col{j+1}"] = "" if val is None else str(val)
            rows.append(row)
        df = pd.DataFrame(rows)
    else:
        df = pd.read_csv(
            path,
            encoding="utf-8-sig",
            dtype={title_col: str, id_col: str},
            keep_default_na=False
        )
    for col in [title_col, id_col]:
        if col not in df.columns:
            raise ValueError(f"Column '{col}' not found in {path}.")
    df[title_col] = df[title_col].astype(str)
    df[id_col] = df[id_col].astype(str)
    return df

def write_output_table(df: pd.DataFrame, output_path: str) -> None:
    """
    Write DataFrame to CSV (UTF-8 with BOM) or XLSX (Unicode).
    Uses file extension to decide format.
    """
    ext = os.path.splitext(output_path.lower())[1]
    if ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        # Excel preserves Unicode; values only (no formulas)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="filtered")
    else:
        # UTF-8 BOM helps Excel display non-ASCII correctly
        df.to_csv(output_path, index=False, encoding="utf-8-sig")

def main():
    parser = argparse.ArgumentParser(
        description="Filter main conferences (ACL, NAACL, AACL/IJCNLP, EMNLP, EACL, LREC, COLING), Findings, TACL, and CL using BibTeX fields only."
    )
    parser.add_argument("csv_path", help="Input CSV or XLSX (first sheet)")
    parser.add_argument("--bibfiles", nargs="+", default=[], help="BibTeX files to parse")
    parser.add_argument("-o", "--output", default="filtered_main_conf.csv", help="Output CSV path")
    parser.add_argument("--title_col", default="title", help="CSV title column")
    parser.add_argument("--id_col", default="id", help="CSV paper id column")
    parser.add_argument("--excel_read", choices=["values", "formulas"], default="values",
                        help="For XLSX, read computed values or raw formulas (openpyxl).")
    args = parser.parse_args()

    df = load_input_table(args.csv_path, args.title_col, args.id_col, excel_read=args.excel_read)
    compiled = compile_patterns(VENUE_PATTERNS)
    bib_index = load_bib_index(args.bibfiles)

    venues: List[Optional[str]] = []
    keep_mask: List[bool] = []
    unmatched_bib = 0

    for _, row in df.iterrows():
        csv_id = str(row.get(args.id_col, "")).strip()

        entry = bib_index.get(csv_id)
        if entry is None:
            unmatched_bib += 1
            venues.append(None)
            keep_mask.append(False)
            continue

        venue = detect_venue(entry, compiled)

        if venue:
            venues.append(venue)
            keep_mask.append(True)
        else:
            venues.append(None)
            keep_mask.append(False)

    df["main_venue_match"] = venues
    filtered = df[pd.Series(keep_mask, index=df.index)]

    counts = filtered["main_venue_match"].value_counts(dropna=False)
    print("Matched venue counts:")
    for v, c in counts.items():
        print(f"  {v}: {c}")

    # print(f"Bib entries not found for {unmatched_bib} CSV ids.")
    # filtered.to_csv(args.output, index=False, encoding="utf-8-sig")
    # print(f"Wrote {len(filtered)} rows to {args.output}")
    write_output_table(filtered, args.output)
    print(f"Wrote {len(filtered)} rows to {args.output}")

if __name__ == "__main__":
    main()